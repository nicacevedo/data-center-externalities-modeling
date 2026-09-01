"""Parse City of Prineville Facebook utility-meter public records.

Observational layer only: ingest, classify, QA, reconcile, and gate.

This module does not equate City meter usage with total Meta campus withdrawal,
campus consumption, wastewater return, groundwater withdrawal, municipal
production, cooling water, or consumptive use.

Raw files under data/raw/city_prineville_public_records_2026/ are immutable.
"""
from __future__ import annotations

import hashlib
import json
import math
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
RAW = ROOT / "data" / "raw" / "city_prineville_public_records_2026"
OUT = ROOT / "data" / "processed" / "city_prineville"
QC = ROOT / "outputs" / "qc"
FIG = ROOT / "outputs" / "city_prineville" / "figures"
META_ANNUAL = ROOT / "data" / "canonical" / "meta_prineville_annual.csv"
OWRD_DIRECT = ROOT / "data" / "processed" / "owrd" / "owrd_meta_direct_monthly_use.csv"

# City explanatory note (source metadata, not a hypothesis).
GAL_PER_100CF = 748
M3_PER_US_GAL = 0.003785411784

SOURCE_ID_METER = "CITY_FB_METER_CONSUMPTION"
SOURCE_ID_BULK = "CITY_FB_BULK_WATER"
SOURCE_ID_EVENTS = "CITY_FB_METER_EVENTS"
SOURCE_ID_NOTE = "CITY_UTILITY_DATA_DESCRIPTION"

PRIMARY_XLSX = "FB Meters and Consumption(2).xlsx"
PRIMARY_CSV = "FB Meters and Consumption(1).csv"
PRIMARY_CSV_DUP = "FB Meters and Consumption.csv"
PRIMARY_TXT = "FB Meters and Consumption.txt"
PRIMARY_PDF = "FB Meters and Consumption.pdf"
BULK_XLSX = "Facebook Bulk Water.xlsx"
EVENTS_XLSX = "Meter Sets, Swaps.xlsx"
NOTE_TXT = "City-of-Prinville-Data-Description.txt"

MONTH_NAMES = {
    "january": 1,
    "february": 2,
    "feburary": 2,  # source spelling
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}

HEADER_USAGE_RE = re.compile(
    r"^(January|February|Feburary|March|April|May|June|July|August|"
    r"September|October|November|December)\s+(\d{4})",
    re.I,
)
TOTAL_RE = re.compile(r"^(\d{4})\s+Total", re.I)

ENTITY_FACEBOOK_DC = "Facebook Data Center"


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def meter_id_to_str(value) -> str:
    """Preserve meter identifiers as strings, including leading zeros.

    Excel may store 0001 as text and 1562600912 as int. Do not coerce to
    int/float. Short 1-4 stored as numeric 1-4 are restored to 0001-0004
    because the native XLSX stores those IDs as four-character text.
    """
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    s = str(value).strip()
    if s.lower() in {"none", "nan", ""}:
        return ""
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    if s in {"1", "2", "3", "4"}:
        return s.zfill(4)
    return s


def parse_numeric(value):
    """Return float or None. Empty is missing, not zero."""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    s = str(value).strip().replace(",", "")
    if s == "" or s.lower() in {"none", "nan", "null"}:
        return None
    try:
        return float(s)
    except ValueError:
        return "malformed"


def usage_gallons(usage_100cf: float) -> float:
    return float(usage_100cf) * GAL_PER_100CF


def usage_m3(usage_100cf: float) -> float:
    return usage_gallons(usage_100cf) * M3_PER_US_GAL


def classify_component(entity: str, rate_code: str) -> dict:
    """Conservative rate-code / entity classification. Labels are not identities."""
    ent = (entity or "").strip()
    rc = (rate_code or "").strip()
    rc_u = rc.upper()

    if ent == "Facebook Trailer City":
        return {
            "component_class": "trailer_city_water",
            "physical_direction": "inflow",
            "semantic_hint": "trailer-city",
            "boundary_status": "unresolved",
            "model_use": "excluded_from_city_service",
            "semantic_note": (
                "Facebook Trailer City account. Not identified as main-campus "
                "operations. Excluded from city_metered_water_service_m3."
            ),
        }
    if ent == "Facebook Warehouse":
        return {
            "component_class": "warehouse_water",
            "physical_direction": "inflow",
            "semantic_hint": "warehouse",
            "boundary_status": "unresolved",
            "model_use": "excluded_from_city_service",
            "semantic_note": (
                "Facebook Warehouse at a different service location. Not identified "
                "as main-campus operations. Excluded from city_metered_water_service_m3."
            ),
        }
    if rc_u.startswith("BULK WATER"):
        return {
            "component_class": "bulk_water",
            "physical_direction": "inflow",
            "semantic_hint": "bulk-hydrant",
            "boundary_status": "unresolved",
            "model_use": "excluded_from_city_service",
            "semantic_note": (
                "Hydrant-meter bulk water organized by City billing convention. "
                "Not identified as ordinary operational cooling water."
            ),
        }
    if "WELL METER FOR SEW" in rc_u:
        return {
            "component_class": "well_meter_for_sew",
            "physical_direction": "unknown",
            "semantic_hint": "unresolved-well-or-sewer-label",
            "boundary_status": "unresolved",
            "model_use": "excluded_from_city_service",
            "semantic_note": (
                "Rate-code label only. Not identified as sewer discharge and not "
                "identified as groundwater withdrawal."
            ),
        }
    if rc_u.startswith("SWR METER") or rc_u == "SWR METER":
        return {
            "component_class": "swr_meter",
            "physical_direction": "unknown",
            "semantic_hint": "sewer-related",
            "boundary_status": "unresolved",
            "model_use": "excluded_from_city_service",
            "semantic_note": (
                "Raw source label is SWR METER. Flow direction is not identified. "
                "Not wastewater return, total discharge, or a consumptive-use offset."
            ),
        }
    if rc_u.startswith("WATER - COMM") or rc_u.startswith("ADD'L WATER"):
        if ent == ENTITY_FACEBOOK_DC:
            kind = "water_comm" if rc_u.startswith("WATER - COMM") else "addl_water"
            return {
                "component_class": kind,
                "physical_direction": "inflow",
                "semantic_hint": "city-metered-service",
                "boundary_status": "provisional",
                "model_use": "included_in_city_service",
                "semantic_note": (
                    "City-metered utility/service water for Facebook Data Center "
                    "records. Not total Meta withdrawal."
                ),
            }
        return {
            "component_class": "other_water_coded",
            "physical_direction": "inflow",
            "semantic_hint": "non-data-center-water-coded",
            "boundary_status": "unresolved",
            "model_use": "excluded_from_city_service",
            "semantic_note": "Water-coded meter on a non-Data-Center entity.",
        }
    return {
        "component_class": "unexpected",
        "physical_direction": "unknown",
        "semantic_hint": "unknown",
        "boundary_status": "unresolved",
        "model_use": "excluded_from_city_service",
        "semantic_note": "Encountered class not in the predeclared list.",
    }


def is_city_service_row(entity: str, rate_code: str) -> bool:
    ent = (entity or "").strip()
    rc = (rate_code or "").strip().upper()
    return ent == ENTITY_FACEBOOK_DC and (
        rc.startswith("WATER - COMM") or rc.startswith("ADD'L WATER")
    )


def _header_month_map(header_row: list) -> tuple[int | None, dict, int | None]:
    """Return (reporting_year, {month: usage_col_idx}, total_col_idx). cols 0-based."""
    year = None
    months = {}
    total_col = None
    for j, val in enumerate(header_row):
        if val is None:
            continue
        text = str(val).replace("\n", " ").strip()
        m = HEADER_USAGE_RE.match(text)
        if m:
            year = int(m.group(2))
            months[MONTH_NAMES[m.group(1).lower()]] = j
            continue
        t = TOTAL_RE.match(text)
        if t:
            year = int(t.group(1))
            total_col = j
    return year, months, total_col


def _iter_sheet_rows(path: Path, sheet: str | None = None):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        yield i, list(row)
    wb.close()


def parse_primary_meter_report(path: Path, source_id: str = SOURCE_ID_METER) -> pd.DataFrame:
    rows_out = []
    current_year = None
    month_cols = {}
    total_col = None
    n_cols = None

    for source_row_id, row in _iter_sheet_rows(path):
        if not any(v is not None and str(v).strip() != "" for v in row):
            continue
        name0 = "" if row[0] is None else str(row[0]).strip()
        if name0 == "Name":
            current_year, month_cols, total_col = _header_month_map(row)
            n_cols = len(row)
            continue
        if current_year is None or not month_cols:
            continue
        entity = name0
        location = "" if row[1] is None else str(row[1]).strip()
        meter_id = meter_id_to_str(row[2] if len(row) > 2 else None)
        if not entity or not meter_id:
            continue
        rate = ""
        if n_cols and len(row) >= n_cols:
            rate = "" if row[n_cols - 1] is None else str(row[n_cols - 1]).strip()
        elif len(row) > 28:
            rate = "" if row[28] is None else str(row[28]).strip()
        annual_raw = None
        if total_col is not None and total_col < len(row):
            annual_raw = parse_numeric(row[total_col])
            if annual_raw == "malformed":
                annual_raw = None
        clf = classify_component(entity, rate)
        for month, usage_idx in sorted(month_cols.items()):
            days_idx = usage_idx + 1
            usage_raw = parse_numeric(row[usage_idx] if usage_idx < len(row) else None)
            days_raw = parse_numeric(row[days_idx] if days_idx < len(row) else None)
            malformed = usage_raw == "malformed" or days_raw == "malformed"
            usage_val = None if usage_raw in (None, "malformed") else float(usage_raw)
            days_val = None if days_raw in (None, "malformed") else float(days_raw)
            rec = {
                "year": int(current_year),
                "month": int(month),
                "source_reporting_year": int(current_year),
                "entity_name": entity,
                "service_location": location,
                "meter_id_raw": meter_id,
                "rate_code_raw": rate,
                "usage_100cf": usage_val,
                "usage_gallons": None if usage_val is None else usage_gallons(usage_val),
                "usage_m3": None if usage_val is None else usage_m3(usage_val),
                "reported_days": days_val,
                "annual_total_raw": annual_raw,
                "source_row_id": f"{path.name}:{source_row_id}",
                "source_id": source_id,
                "observation_status": "malformed" if malformed else "",
                "partial_period_flag": False,
                **clf,
            }
            rows_out.append(rec)
    long = pd.DataFrame(rows_out)
    if long.empty:
        return long
    return _assign_observation_status(long)


def parse_primary_delimited(path: Path, source_id: str = "CITY_FB_METER_CONSUMPTION_MIRROR") -> pd.DataFrame:
    """Parse CSV/TXT semicolon export of the same annual-block report."""
    df = pd.read_csv(path, sep=";", dtype=str, encoding="utf-8-sig", header=None)
    rows_out = []
    current_year = None
    month_cols = {}
    total_col = None
    for i, row in df.iterrows():
        vals = [None if pd.isna(v) else v for v in row.tolist()]
        name0 = "" if vals[0] is None else str(vals[0]).strip()
        if name0 == "Name":
            current_year, month_cols, total_col = _header_month_map(vals)
            continue
        if current_year is None or not month_cols:
            continue
        entity = name0
        location = "" if vals[1] is None else str(vals[1]).strip()
        meter_id = meter_id_to_str(vals[2] if len(vals) > 2 else None)
        if not entity or not meter_id:
            continue
        rate = "" if vals[-1] is None else str(vals[-1]).strip()
        annual_raw = parse_numeric(vals[total_col] if total_col is not None else None)
        if annual_raw == "malformed":
            annual_raw = None
        clf = classify_component(entity, rate)
        for month, usage_idx in sorted(month_cols.items()):
            usage_raw = parse_numeric(vals[usage_idx] if usage_idx < len(vals) else None)
            days_raw = parse_numeric(vals[usage_idx + 1] if usage_idx + 1 < len(vals) else None)
            malformed = usage_raw == "malformed" or days_raw == "malformed"
            usage_val = None if usage_raw in (None, "malformed") else float(usage_raw)
            days_val = None if days_raw in (None, "malformed") else float(days_raw)
            rows_out.append(
                {
                    "year": int(current_year),
                    "month": int(month),
                    "meter_id_raw": meter_id,
                    "entity_name": entity,
                    "usage_100cf": usage_val,
                    "reported_days": days_val,
                    "annual_total_raw": annual_raw,
                    "rate_code_raw": rate,
                    "source_row_id": f"{path.name}:{i + 1}",
                    "source_id": source_id,
                    "malformed": malformed,
                    **clf,
                }
            )
    return pd.DataFrame(rows_out)


def _assign_observation_status(long: pd.DataFrame) -> pd.DataFrame:
    """Distinguish observed zero, missing, structurally unavailable, not-yet-observed."""
    out = long.copy()
    dc = out[out["entity_name"].eq(ENTITY_FACEBOOK_DC)]
    observed = dc[(dc["reported_days"].fillna(0) > 0)]
    if observed.empty:
        last_year, last_month = None, None
    else:
        last_year = int(observed["year"].max())
        last_month = int(observed.loc[observed["year"].eq(last_year), "month"].max())

    # First/last month with positive read-days in each meter-year (for gap vs not-yet-in-service).
    first_active = {}
    last_active = {}
    for (meter, year), g in out.groupby(["meter_id_raw", "year"]):
        active = g[pd.to_numeric(g["reported_days"], errors="coerce").fillna(0) > 0]
        if active.empty:
            continue
        first_active[(meter, int(year))] = int(active["month"].min())
        last_active[(meter, int(year))] = int(active["month"].max())

    statuses = []
    partials = []
    typical = {1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30, 7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31}
    for r in out.itertuples(index=False):
        if r.observation_status == "malformed":
            statuses.append("malformed")
            partials.append(False)
            continue
        days = r.reported_days
        usage = r.usage_100cf
        y, m = int(r.year), int(r.month)
        days_n = 0.0 if days is None or (isinstance(days, float) and math.isnan(days)) else float(days)
        usage_n = 0.0 if usage is None or (isinstance(usage, float) and math.isnan(usage)) else float(usage)
        after_last = (
            last_year is not None
            and (y > last_year or (y == last_year and m > last_month))
        )
        if days_n > 0:
            if after_last:
                statuses.append("source_anomaly")
            elif usage is None:
                statuses.append("missing")
            elif usage_n == 0:
                statuses.append("observed_zero")
            else:
                statuses.append("observed")
            partials.append(days_n > 0 and days_n < typical.get(m, 28) - 1)
            continue
        # days == 0 or missing
        if after_last and usage_n == 0:
            statuses.append("not_observed_yet")
            partials.append(True)
            continue
        if usage_n != 0 and days_n == 0:
            statuses.append("source_anomaly")
            partials.append(False)
            continue
        fa = first_active.get((r.meter_id_raw, y))
        la = last_active.get((r.meter_id_raw, y))
        if fa is not None and la is not None and fa < m < la:
            statuses.append("missing")
            partials.append(True)
            continue
        if fa is not None and m < fa:
            statuses.append("structurally_unavailable")
            partials.append(False)
            continue
        if la is not None and m > la and not after_last:
            statuses.append("structurally_unavailable")
            partials.append(False)
            continue
        if usage is None:
            statuses.append("missing")
        else:
            statuses.append("structurally_unavailable")
        partials.append(False)
    out["observation_status"] = statuses
    out["partial_period_flag"] = partials
    out.attrs["latest_observed_year"] = last_year
    out.attrs["latest_observed_month"] = last_month
    return out


def compare_meter_representations(xlsx_long: pd.DataFrame, other: pd.DataFrame) -> dict:
    keys = ["year", "meter_id_raw", "month"]
    a = xlsx_long[keys + ["usage_100cf", "reported_days"]].copy()
    b = other[keys + ["usage_100cf", "reported_days"]].copy()
    a["_u"] = pd.to_numeric(a["usage_100cf"], errors="coerce")
    a["_d"] = pd.to_numeric(a["reported_days"], errors="coerce")
    b["_u"] = pd.to_numeric(b["usage_100cf"], errors="coerce")
    b["_d"] = pd.to_numeric(b["reported_days"], errors="coerce")
    merged = a.merge(b, on=keys, how="outer", suffixes=("_xlsx", "_other"), indicator=True)
    both = merged["_merge"].eq("both")
    u_diff = both & ((merged["_u_xlsx"] - merged["_u_other"]).abs() > 1e-9)
    d_diff = both & ((merged["_d_xlsx"] - merged["_d_other"]).abs() > 1e-9)
    return {
        "xlsx_rows": int(len(a)),
        "other_rows": int(len(b)),
        "matched": int(both.sum()),
        "xlsx_only": int(merged["_merge"].eq("left_only").sum()),
        "other_only": int(merged["_merge"].eq("right_only").sum()),
        "usage_mismatches": int(u_diff.sum()),
        "days_mismatches": int(d_diff.sum()),
        "materially_equivalent": bool(
            (merged["_merge"].eq("both").all() or (int(merged["_merge"].eq("left_only").sum()) == 0 and int(merged["_merge"].eq("right_only").sum()) == 0))
            and int(u_diff.sum()) == 0
            and int(d_diff.sum()) == 0
            and int(len(a)) == int(len(b))
        ),
    }


def pdf_evidence_tokens(pdf_path: Path) -> dict:
    try:
        import pymupdf as fitz
    except ImportError:
        try:
            import fitz
        except ImportError:
            return {"available": False, "tokens_found": [], "tokens_missing": []}
    doc = fitz.open(pdf_path)
    text = "\n".join(page.get_text() for page in doc)
    n_pages = doc.page_count
    n_chars = len(text)
    doc.close()
    tokens = [
        "Facebook Data Center",
        "Facebook Trailer City",
        "Facebook Warehouse",
        "WELL METER FOR SEW",
        "SWR METER",
        "WATER - COMM",
        "0001",
        "0002",
        "0003",
        "0004",
        "2012",
        "2026",
        "735 SW Connect Way",
    ]
    found = [t for t in tokens if t in text]
    missing = [t for t in tokens if t not in text]
    return {
        "available": True,
        "n_chars": n_chars,
        "n_pages": n_pages,
        "tokens_found": found,
        "tokens_missing": missing,
        "likely_image_or_vector_table": n_chars < 5000,
        "sufficient_evidence_mirror": (len(missing) <= 2) or (n_chars < 5000 and len(found) >= 6),
        "note": (
            "PDF is a delivery/evidence copy, not an independent observation. "
            "Extracted text is sparse relative to file size."
        ),
    }


def parse_bulk_water(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        if not any(v is not None for v in row):
            continue
        bill_date = pd.to_datetime(row[0], errors="coerce")
        rate = "" if row[1] is None else str(row[1]).strip()
        cons = parse_numeric(row[2] if len(row) > 2 else None)
        days = parse_numeric(row[3] if len(row) > 3 else None)
        malformed = cons == "malformed" or days == "malformed"
        usage_val = None if cons in (None, "malformed") else float(cons)
        days_val = None if days in (None, "malformed") else float(days)
        clf = classify_component(ENTITY_FACEBOOK_DC, rate or "BULK WATER")
        if bill_date is pd.NaT or pd.isna(bill_date):
            bill_year = None
            bill_month = None
            bill_iso = None
        else:
            bill_year = int(bill_date.year)
            bill_month = int(bill_date.month)
            bill_iso = bill_date.strftime("%Y-%m-%d")
        if malformed:
            status = "malformed"
        elif usage_val is None:
            status = "missing"
        elif (days_val or 0) > 0 and usage_val == 0:
            status = "observed_zero"
        elif (days_val or 0) > 0:
            status = "observed"
        else:
            status = "source_anomaly"
        rows.append(
            {
                "bill_date": bill_iso,
                "bill_year": bill_year,
                "bill_month": bill_month,
                "rate_code_raw": rate,
                "usage_100cf": usage_val,
                "usage_gallons": None if usage_val is None else usage_gallons(usage_val),
                "usage_m3": None if usage_val is None else usage_m3(usage_val),
                "reported_days": days_val,
                "source_row_id": f"{path.name}:{i}",
                "source_id": SOURCE_ID_BULK,
                "observation_status": status,
                "time_basis": "city_billing_convention",
                "consumption_month_claimed": False,
                **clf,
            }
        )
    wb.close()
    return pd.DataFrame(rows)


def parse_meter_events(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        if not any(v is not None for v in row):
            continue
        action = "" if row[0] is None else str(row[0]).rstrip()
        old_raw = meter_id_to_str(row[1] if len(row) > 1 else None)
        new_raw = meter_id_to_str(row[2] if len(row) > 2 else None)
        dt = pd.to_datetime(row[3] if len(row) > 3 else None, errors="coerce")
        rows.append(
            {
                "action": action,
                "old_meter_raw": old_raw,
                "new_meter_raw": new_raw,
                "event_date": None if pd.isna(dt) else dt.strftime("%Y-%m-%d"),
                "event_year": None if pd.isna(dt) else int(dt.year),
                "event_month": None if pd.isna(dt) else int(dt.month),
                "source_row_id": f"{path.name}:{i}",
                "source_id": SOURCE_ID_EVENTS,
                "pre_2018_history_complete": False,
            }
        )
    wb.close()
    return pd.DataFrame(rows)


def _leading_1_candidate(meter_id: str) -> str | None:
    if meter_id and meter_id.startswith("1") and len(meter_id) >= 2 and meter_id[1] != "0":
        return meter_id[1:]
    return None


def build_lineage_audit(events: pd.DataFrame, meters: pd.DataFrame) -> pd.DataFrame:
    consumption_ids = set(meters["meter_id_raw"].astype(str).unique())
    rows = []
    for r in events.itertuples(index=False):
        for role, raw in (("old", r.old_meter_raw), ("new", r.new_meter_raw)):
            if not raw:
                continue
            exact = raw in consumption_ids
            cand = _leading_1_candidate(raw)
            inferred = bool(cand and cand in consumption_ids)
            if exact and inferred and cand != raw:
                match_type = "ambiguous"
                accepted = False
                confidence = "low"
                evidence = "raw ID and leading-1 candidate both in consumption report"
            elif exact:
                match_type = "exact"
                accepted = True
                confidence = "high"
                evidence = "event meter ID equals a consumption-report meter ID"
                cand = raw
            elif inferred:
                match_type = "inferred_leading_1"
                accepted = False
                confidence = "medium"
                evidence = (
                    f"event ID {raw} matches consumption ID {cand} after dropping one leading 1; "
                    "not silently normalized"
                )
            else:
                match_type = "unmatched"
                accepted = False
                confidence = "none"
                evidence = "no exact or leading-1 match in consumption report"
                cand = cand or ""
            rows.append(
                {
                    "source_row_id": r.source_row_id,
                    "action": r.action,
                    "event_date": r.event_date,
                    "event_meter_role": role,
                    "event_meter_id_raw": raw,
                    "candidate_consumption_meter_id": cand or "",
                    "match_rule": "exact_or_drop_leading_1",
                    "match_type": match_type,
                    "evidence": evidence,
                    "confidence": confidence,
                    "accepted_for_lineage_boolean": accepted,
                }
            )
    return pd.DataFrame(rows)


def _active_observation(status: str) -> bool:
    return status in {"observed", "observed_zero"}


def build_qa(
    hashes: pd.DataFrame,
    equiv: dict,
    txt_vs_csv: dict,
    pdf_info: dict,
    meters: pd.DataFrame,
    bulk: pd.DataFrame,
    events: pd.DataFrame,
    lineage: pd.DataFrame,
    meta_hash_before: str,
    meta_hash_after: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    exceptions = []
    checks = []

    def add_check(name, status, detail, severity="info"):
        checks.append(
            {
                "check_id": name,
                "status": status,
                "severity": severity,
                "detail": detail,
            }
        )

    def add_exc(check_id, severity, status, detail, **extra):
        row = {
            "check_id": check_id,
            "severity": severity,
            "status": status,
            "detail": detail,
        }
        row.update(extra)
        exceptions.append(row)

    add_check("source_file_hashes", "PASS", f"{len(hashes)} artifacts hashed")

    mat = equiv.get("materially_equivalent", False)
    add_check(
        "primary_report_xlsx_csv_equivalence",
        "PASS" if mat else "FAIL",
        json.dumps({k: v for k, v in equiv.items() if k != "materially_equivalent"}),
        "blocking" if not mat else "info",
    )
    add_check(
        "primary_report_csv_txt_newline_only",
        "PASS" if txt_vs_csv.get("equivalent_after_newline_normalize") else "FAIL",
        json.dumps(txt_vs_csv),
        "blocking" if not txt_vs_csv.get("equivalent_after_newline_normalize") else "info",
    )
    pdf_ok = pdf_info.get("sufficient_evidence_mirror", False)
    add_check(
        "primary_report_pdf_evidence_mirror",
        "PASS" if pdf_ok else "WARN",
        json.dumps(pdf_info, default=str),
        "warning" if not pdf_ok else "info",
    )

    # unit conversion spot checks encoded as checks (values tested in pytest too)
    add_check(
        "unit_conversion",
        "PASS",
        f"1 unit = {GAL_PER_100CF} gal = {usage_m3(1)} m3; 0 units = 0",
    )

    leading = meters[meters["meter_id_raw"].isin(["0001", "0002", "0003", "0004"])]
    stripped = meters[meters["meter_id_raw"].isin(["1", "2", "3", "4"])]
    leading_ok = (len(stripped) == 0) and (not leading.empty)
    add_check(
        "meter_id_leading_zeros",
        "PASS" if leading_ok else "FAIL",
        f"padded IDs n={len(leading)}; stripped numeric IDs n={len(stripped)}",
        "blocking" if not leading_ok else "info",
    )

    years = sorted(meters["year"].unique())
    add_check(
        "repeated_annual_blocks",
        "PASS" if years == list(range(int(min(years)), int(max(years)) + 1)) else "WARN",
        f"years={[int(y) for y in years]}",
    )
    bad_ym = meters[(meters["month"] < 1) | (meters["month"] > 12) | (meters["year"] < 2000)]
    add_check(
        "valid_year_month",
        "PASS" if bad_ym.empty else "FAIL",
        f"invalid rows={len(bad_ym)}",
        "blocking" if not bad_ym.empty else "info",
    )

    dup_src = meters.duplicated(["source_row_id", "month"], keep=False)
    add_check(
        "source_row_uniqueness",
        "PASS" if not dup_src.any() else "FAIL",
        f"duplicate source_row_id+month={int(dup_src.sum())}",
        "blocking" if dup_src.any() else "info",
    )

    mm = meters.duplicated(["year", "month", "meter_id_raw", "entity_name"], keep=False)
    add_check(
        "meter_month_duplicates",
        "PASS" if not mm.any() else "FAIL",
        f"duplicate meter-month rows={int(mm.sum())}",
        "blocking" if mm.any() else "info",
    )

    # annual total vs monthly sum, observed+zero+structurally_unavailable numeric zeros
    annual_mismatch = 0
    for (meter, year, src_row), g in meters.groupby(["meter_id_raw", "year", "source_row_id"]):
        reported = g["annual_total_raw"].iloc[0]
        # Source annual total includes all twelve listed cells, including future 0/0.
        monthly_sum = pd.to_numeric(g["usage_100cf"], errors="coerce").fillna(0).sum()
        if reported is None or (isinstance(reported, float) and math.isnan(reported)):
            continue
        if abs(monthly_sum - float(reported)) > 1e-6:
            annual_mismatch += 1
            add_exc(
                "annual_total_mismatch",
                "warning",
                "source_anomaly",
                f"meter={meter} year={year} monthly_sum={monthly_sum} annual_total_raw={reported}",
                meter_id_raw=meter,
                year=year,
            )
    add_check(
        "monthly_sum_vs_annual_total",
        "PASS" if annual_mismatch == 0 else "WARN",
        f"mismatched meter-years={annual_mismatch}",
        "warning" if annual_mismatch else "info",
    )

    nz_zero_days = meters[
        (pd.to_numeric(meters["usage_100cf"], errors="coerce").fillna(0) != 0)
        & (pd.to_numeric(meters["reported_days"], errors="coerce").fillna(0) == 0)
    ]
    for r in nz_zero_days.itertuples(index=False):
        add_exc(
            "nonzero_usage_zero_days",
            "warning",
            "source_anomaly",
            f"{r.meter_id_raw} {r.year}-{r.month:02d} usage={r.usage_100cf}",
            meter_id_raw=r.meter_id_raw,
            year=r.year,
            month=r.month,
        )
    add_check(
        "nonzero_usage_zero_read_days",
        "PASS" if nz_zero_days.empty else "WARN",
        f"n={len(nz_zero_days)}",
        "warning" if not nz_zero_days.empty else "info",
    )

    z_pos_days = meters[
        (pd.to_numeric(meters["usage_100cf"], errors="coerce").fillna(0) == 0)
        & (pd.to_numeric(meters["reported_days"], errors="coerce").fillna(0) > 0)
        & meters["observation_status"].eq("observed_zero")
    ]
    add_check(
        "zero_usage_positive_read_days",
        "PASS",
        f"observed_zero n={len(z_pos_days)} (valid observed zeros, not missing)",
    )

    future = meters[meters["observation_status"].eq("not_observed_yet")]
    add_check(
        "unobserved_2026_future_periods",
        "PASS" if not future.empty else "WARN",
        f"not_observed_yet n={len(future)}; latest observed="
        f"{meters.attrs.get('latest_observed_year')}-{meters.attrs.get('latest_observed_month')}",
        "warning" if future.empty else "info",
    )
    if (future["usage_100cf"].fillna(0) != 0).any():
        add_check("future_nonzero", "FAIL", "unobserved future cells have nonzero usage", "blocking")

    # bulk duplicate bill dates
    if not bulk.empty:
        dup_dates = bulk.groupby("bill_date").size()
        n_dup = int((dup_dates > 1).sum())
        add_check(
            "duplicate_bulk_bill_dates",
            "WARN" if n_dup else "PASS",
            f"dates with >1 row={n_dup}; preserved as separate hydrant-meter rows, summed only in monthly aggregates",
            "warning" if n_dup else "info",
        )
        for d, n in dup_dates[dup_dates > 1].items():
            add_exc(
                "duplicate_bulk_bill_dates",
                "warning",
                "expected_source_structure",
                f"bill_date={d} n_rows={int(n)}",
                bill_date=d,
            )

    cons_ids = set(meters["meter_id_raw"].astype(str))
    event_ids = set()
    for col in ("old_meter_raw", "new_meter_raw"):
        event_ids.update(x for x in events[col].astype(str) if x and x != "nan")
    event_not_cons = sorted(event_ids - cons_ids)
    cons_not_event = sorted(cons_ids - event_ids)
    add_check(
        "event_ids_absent_from_consumption",
        "WARN",
        f"n={len(event_not_cons)} (includes leading-1 event IDs and unmatched); not a parser error",
        "warning",
    )
    for mid in event_not_cons:
        add_exc(
            "event_ids_absent_from_consumption",
            "warning",
            "expected_source_limitation" if mid.startswith("1") else "warning",
            f"event meter {mid} not in consumption report as raw ID",
            meter_id_raw=mid,
        )
    add_check(
        "consumption_ids_absent_from_events",
        "WARN",
        f"n={len(cons_not_event)}; pre-2018 lifecycle is incomplete by City note",
        "expected_limitation",
    )

    # usage outside lifecycle: consumption before set / after pull using accepted exact matches only
    set_dates = {}
    pull_dates = {}
    for r in events.itertuples(index=False):
        act = (r.action or "").lower()
        if "set" in act and r.new_meter_raw:
            set_dates.setdefault(r.new_meter_raw, []).append(r.event_date)
        if "pull" in act and r.old_meter_raw:
            pull_dates.setdefault(r.old_meter_raw, []).append(r.event_date)
    n_before_set = 0
    n_after_pull = 0
    for r in meters.itertuples(index=False):
        if r.observation_status not in {"observed", "observed_zero"}:
            continue
        period = f"{int(r.year):04d}-{int(r.month):02d}-01"
        for sd in set_dates.get(r.meter_id_raw, []):
            if sd and period < sd[:7] + "-01" and f"{int(r.year):04d}-{int(r.month):02d}" < sd[:7]:
                # consumption month strictly before event month
                if (int(r.year), int(r.month)) < (int(sd[:4]), int(sd[5:7])):
                    n_before_set += 1
                    add_exc(
                        "usage_before_set_date",
                        "warning",
                        "source_anomaly",
                        f"{r.meter_id_raw} {r.year}-{int(r.month):02d} before set {sd}",
                        meter_id_raw=r.meter_id_raw,
                        year=r.year,
                        month=r.month,
                    )
                    break
        for pd_ in pull_dates.get(r.meter_id_raw, []):
            if pd_ and (int(r.year), int(r.month)) > (int(pd_[:4]), int(pd_[5:7])):
                n_after_pull += 1
                add_exc(
                    "usage_after_pull_date",
                    "warning",
                    "source_anomaly",
                    f"{r.meter_id_raw} {r.year}-{int(r.month):02d} after pull {pd_}",
                    meter_id_raw=r.meter_id_raw,
                    year=r.year,
                    month=r.month,
                )
                break
    add_check(
        "usage_outside_lifecycle_intervals",
        "WARN" if (n_before_set or n_after_pull) else "PASS",
        f"before_set={n_before_set} after_pull={n_after_pull}",
        "warning" if (n_before_set or n_after_pull) else "info",
    )

    anomaly_157 = meters[
        meters["meter_id_raw"].eq("1573376176")
        & meters["observation_status"].isin(["observed", "observed_zero"])
        & ((meters["year"] < 2026) | ((meters["year"] == 2026) & (meters["month"] < 2)))
    ]
    add_check(
        "anomalous_event_date_1573376176",
        "WARN",
        "Meter 1573376176 has a City set date 2026-02-15 but appears in consumption before that date. Source not corrected.",
        "warning",
    )
    if not anomaly_157.empty:
        add_exc(
            "anomalous_event_date_1573376176",
            "warning",
            "source_anomaly",
            f"n_observed_months_before_set={len(anomaly_157)}",
            meter_id_raw="1573376176",
        )

    ents = sorted(meters["entity_name"].unique())
    unexpected_ent = [e for e in ents if e not in {ENTITY_FACEBOOK_DC, "Facebook Trailer City", "Facebook Warehouse"}]
    add_check(
        "unexpected_entities",
        "FAIL" if unexpected_ent else "PASS",
        f"entities={ents}",
        "blocking" if unexpected_ent else "info",
    )
    rates = sorted(meters["rate_code_raw"].unique())
    expected_prefix = ("WATER - COMM", "ADD'L WATER", "SWR METER", "WELL METER FOR SEW")
    unexpected_rc = [r for r in rates if not any(r.startswith(p) for p in expected_prefix)]
    add_check(
        "unexpected_rate_codes",
        "FAIL" if unexpected_rc else "PASS",
        f"rate_codes={rates}",
        "blocking" if unexpected_rc else "info",
    )
    mal = meters[meters["observation_status"].eq("malformed")]
    add_check(
        "malformed_numeric_values",
        "FAIL" if not mal.empty else "PASS",
        f"n={len(mal)}",
        "blocking" if not mal.empty else "info",
    )

    add_check(
        "canonical_meta_annual_untouched",
        "PASS" if meta_hash_before == meta_hash_after else "FAIL",
        f"sha256={meta_hash_after}",
        "blocking" if meta_hash_before != meta_hash_after else "info",
    )

    add_check(
        "pre_2018_lifecycle_incomplete",
        "PASS",
        "City states meter set/swap/inactivation records are incomplete before ~2018. Absence of an event is not proof the meter did not exist.",
        "expected_limitation",
    )

    qa = pd.DataFrame(checks)
    exc = pd.DataFrame(exceptions) if exceptions else pd.DataFrame(
        columns=["check_id", "severity", "status", "detail"]
    )
    if "meter_id_raw" in exc.columns:
        exc["meter_id_raw"] = exc["meter_id_raw"].map(lambda x: "" if pd.isna(x) else str(x))
    return qa, exc


def parent_submeter_investigation(meters: pd.DataFrame, events: pd.DataFrame) -> dict:
    """Look for parent/child or replacement-overlap double counting in City-service meters."""
    svc = meters[meters.apply(lambda r: is_city_service_row(r.entity_name, r.rate_code_raw), axis=1)]
    svc = svc[svc["observation_status"].isin(["observed", "observed_zero"])]
    # Simultaneous distinct service meters in a month is expected for parallel services.
    n_sim = (
        svc.groupby(["year", "month"])["meter_id_raw"].nunique()
        if not svc.empty
        else pd.Series(dtype=int)
    )
    # Swap-month overlap: old consumption ID and new ID both observed in event month.
    overlap_rows = []
    cons_ids = set(meters["meter_id_raw"].astype(str))
    for r in events.itertuples(index=False):
        if "swap" not in (r.action or "").lower():
            continue
        old_c = r.old_meter_raw if r.old_meter_raw in cons_ids else (_leading_1_candidate(r.old_meter_raw) or "")
        new_c = r.new_meter_raw
        if not old_c or not new_c or not r.event_date:
            continue
        y, m = int(r.event_date[:4]), int(r.event_date[5:7])
        old_u = svc[(svc.meter_id_raw.eq(old_c)) & (svc.year.eq(y)) & (svc.month.eq(m))]
        new_u = svc[(svc.meter_id_raw.eq(new_c)) & (svc.year.eq(y)) & (svc.month.eq(m))]
        if not old_u.empty and not new_u.empty:
            overlap_rows.append(
                {
                    "event": r.source_row_id,
                    "old": old_c,
                    "new": new_c,
                    "year": y,
                    "month": m,
                    "old_m3": float(old_u["usage_m3"].sum()),
                    "new_m3": float(new_u["usage_m3"].sum()),
                }
            )
    identified_double_count = False
    note = (
        "No parent/submeter field exists in the City files. Multiple WATER-COMM / ADD'L WATER "
        "meters at 735 SW Connect Way have simultaneous usage; rate codes indicate parallel "
        "service sizes, not nested submeters. "
    )
    if overlap_rows:
        note += (
            f"{len(overlap_rows)} swap-month(s) have both predecessor and successor usage; "
            "those months may overstate the service aggregate by an unresolved replacement overlap. "
            "Not treated as demonstrated parent/submeter double counting."
        )
    else:
        note += "No swap-month predecessor/successor overlap was found in City-service meters."
    return {
        "identified_parent_submeter_double_counting": identified_double_count,
        "max_simultaneous_service_meters": int(n_sim.max()) if len(n_sim) else 0,
        "swap_month_overlaps": overlap_rows,
        "boundary_status": "provisional",
        "note": note,
    }


def components_monthly(meters: pd.DataFrame, bulk: pd.DataFrame) -> pd.DataFrame:
    obs = meters[meters["observation_status"].isin(["observed", "observed_zero"])].copy()
    years = range(int(obs["year"].min()), int(obs["year"].max()) + 1)
    months = range(1, 13)
    idx = pd.MultiIndex.from_product([years, months], names=["year", "month"])
    out = pd.DataFrame(index=idx).reset_index()

    def sum_m3(mask) -> pd.Series:
        g = obs.loc[mask].groupby(["year", "month"])["usage_m3"].sum()
        return out.set_index(["year", "month"]).index.map(lambda k: g.get(k, np.nan))

    svc_mask = obs.apply(lambda r: is_city_service_row(r.entity_name, r.rate_code_raw), axis=1)
    out["city_metered_water_service_m3"] = sum_m3(svc_mask)
    out["swr_meter_volume_m3"] = sum_m3(obs["component_class"].eq("swr_meter"))
    out["well_meter_for_sew_volume_m3"] = sum_m3(obs["component_class"].eq("well_meter_for_sew"))
    out["trailer_city_water_m3"] = sum_m3(obs["component_class"].eq("trailer_city_water"))
    out["warehouse_water_m3"] = sum_m3(obs["component_class"].eq("warehouse_water"))

    # Coverage: NaN where no observed City-service records that month (including not_observed_yet).
    svc_obs = obs[svc_mask]
    covered = set(zip(svc_obs["year"], svc_obs["month"]))
    out.loc[
        ~out.apply(lambda r: (int(r.year), int(r.month)) in covered, axis=1),
        "city_metered_water_service_m3",
    ] = np.nan

    bulk_obs = bulk[bulk["observation_status"].isin(["observed", "observed_zero"])]
    bulk_m = bulk_obs.groupby(["bill_year", "bill_month"])["usage_m3"].sum()
    out["bulk_water_bill_month_m3"] = [
        bulk_m.get((int(y), int(m)), np.nan) if (int(y), int(m)) in bulk_m.index else (
            bulk_m.get((int(y), int(m)), np.nan)
        )
        for y, m in zip(out["year"], out["month"])
    ]
    # If bulk has any row for that bill month, use sum (including zeros); else NaN.
    bulk_months = set(zip(bulk_obs["bill_year"], bulk_obs["bill_month"]))
    out["bulk_water_bill_month_m3"] = [
        float(bulk_m.get((int(y), int(m)), 0.0)) if (int(y), int(m)) in bulk_months else np.nan
        for y, m in zip(out["year"], out["month"])
    ]
    out["city_metered_water_service_note"] = (
        "City-metered utility/service water for Facebook Data Center WATER-COMM + ADD'L WATER records. "
        "Not total Meta withdrawal."
    )
    out["bulk_water_time_basis"] = "bill_month_not_consumption_month"
    return out


def coverage_table(meters: pd.DataFrame, bulk: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for (src, comp, y, m), g in meters.groupby(
        ["source_id", "component_class", "year", "month"]
    ):
        statuses = sorted(g["observation_status"].unique())
        rows.append(
            {
                "source_id": src,
                "component": comp,
                "year": int(y),
                "month": int(m),
                "n_meter_rows": int(len(g)),
                "n_observed": int(g["observation_status"].isin(["observed", "observed_zero"]).sum()),
                "n_not_observed_yet": int(g["observation_status"].eq("not_observed_yet").sum()),
                "observation_statuses": "|".join(statuses),
                "usage_m3_sum_observed": float(
                    g.loc[g["observation_status"].isin(["observed", "observed_zero"]), "usage_m3"].sum()
                ),
            }
        )
    if not bulk.empty:
        for (y, m), g in bulk.groupby(["bill_year", "bill_month"]):
            rows.append(
                {
                    "source_id": SOURCE_ID_BULK,
                    "component": "bulk_water",
                    "year": int(y),
                    "month": int(m),
                    "n_meter_rows": int(len(g)),
                    "n_observed": int(g["observation_status"].isin(["observed", "observed_zero"]).sum()),
                    "n_not_observed_yet": 0,
                    "observation_statuses": "|".join(sorted(g["observation_status"].unique())),
                    "usage_m3_sum_observed": float(
                        g.loc[g["observation_status"].isin(["observed", "observed_zero"]), "usage_m3"].sum()
                    ),
                }
            )
    return pd.DataFrame(rows).sort_values(["source_id", "component", "year", "month"])


def meta_reconciliation(components: pd.DataFrame, meta: pd.DataFrame) -> pd.DataFrame:
    """Year-by-year boundary reconciliation. Not model fit and not a chosen physical total."""
    annual = components.groupby("year", as_index=False).agg(
        city_metered_water_service_m3=("city_metered_water_service_m3", "sum"),
        n_service_months=("city_metered_water_service_m3", lambda s: int(s.notna().sum())),
        bulk_water_bill_month_m3=("bulk_water_bill_month_m3", "sum"),
        n_bulk_months=("bulk_water_bill_month_m3", lambda s: int(s.notna().sum())),
        swr_meter_volume_m3=("swr_meter_volume_m3", "sum"),
        well_meter_for_sew_volume_m3=("well_meter_for_sew_volume_m3", "sum"),
        trailer_city_water_m3=("trailer_city_water_m3", "sum"),
        warehouse_water_m3=("warehouse_water_m3", "sum"),
    )
    # Sum of NaNs in pandas skipna defaults to 0 for empty — restore NaN if no months.
    for col, ncl in [
        ("city_metered_water_service_m3", "n_service_months"),
        ("bulk_water_bill_month_m3", "n_bulk_months"),
    ]:
        annual.loc[annual[ncl] == 0, col] = np.nan

    m = meta[["year", "water_withdrawal_m3_reported"]].copy()
    m["year"] = m["year"].astype(int)
    out = m.merge(annual, on="year", how="outer").sort_values("year")
    out = out.rename(columns={"water_withdrawal_m3_reported": "meta_annual_withdrawal_m3"})
    out["diagnostic_service_plus_bulk_m3"] = out["city_metered_water_service_m3"] + out["bulk_water_bill_month_m3"]
    out["residual_meta_minus_service_m3"] = out["meta_annual_withdrawal_m3"] - out["city_metered_water_service_m3"]
    out["residual_meta_minus_service_plus_bulk_m3"] = (
        out["meta_annual_withdrawal_m3"] - out["diagnostic_service_plus_bulk_m3"]
    )
    out["city_service_minus_meta_m3"] = out["city_metered_water_service_m3"] - out["meta_annual_withdrawal_m3"]
    out["city_service_plus_bulk_minus_meta_m3"] = (
        out["diagnostic_service_plus_bulk_m3"] - out["meta_annual_withdrawal_m3"]
    )
    with np.errstate(divide="ignore", invalid="ignore"):
        out["pct_residual_service"] = np.where(
            out["meta_annual_withdrawal_m3"].abs() > 0,
            100.0 * out["residual_meta_minus_service_m3"] / out["meta_annual_withdrawal_m3"],
            np.nan,
        )
        out["pct_residual_service_plus_bulk"] = np.where(
            out["meta_annual_withdrawal_m3"].abs() > 0,
            100.0 * out["residual_meta_minus_service_plus_bulk_m3"] / out["meta_annual_withdrawal_m3"],
            np.nan,
        )
        out["city_service_share_of_meta"] = np.where(
            out["meta_annual_withdrawal_m3"].abs() > 0,
            out["city_metered_water_service_m3"] / out["meta_annual_withdrawal_m3"],
            np.nan,
        )
        out["city_service_plus_bulk_share_of_meta"] = np.where(
            out["meta_annual_withdrawal_m3"].abs() > 0,
            out["diagnostic_service_plus_bulk_m3"] / out["meta_annual_withdrawal_m3"],
            np.nan,
        )
        out["city_service_share_of_meta_pct"] = 100.0 * out["city_service_share_of_meta"]
        out["city_service_plus_bulk_share_of_meta_pct"] = 100.0 * out["city_service_plus_bulk_share_of_meta"]
    out["combination_label"] = "diagnostic_service_plus_bulk_allocated_by_observed_bill_year_not_canonical_total"
    out["note"] = (
        "service + bulk (bulk allocated by observed bill year) is an accounting-date "
        "reconciliation diagnostic, not proof of same-year physical campus withdrawal "
        "and not a water-balance closure. Shares and residuals are descriptive only. "
        "A close year is not an accounting identity."
    )
    return out


def owrd_well_meter_compare(meters: pd.DataFrame, owrd_path: Path) -> pd.DataFrame:
    well = meters[
        meters["component_class"].eq("well_meter_for_sew")
        & meters["observation_status"].isin(["observed", "observed_zero"])
    ]
    city_m = well.groupby(["year", "month"], as_index=False)["usage_m3"].sum()
    city_m = city_m.rename(columns={"year": "calendar_year", "month": "calendar_month", "usage_m3": "well_meter_for_sew_m3"})
    if not owrd_path.exists():
        city_m["owrd_direct_pod_m3"] = np.nan
        city_m["note"] = "OWRD direct monthly file missing"
        return city_m
    owrd = pd.read_csv(owrd_path)
    owrd["calendar_year"] = owrd["calendar_year"].astype(int)
    owrd["calendar_month"] = pd.to_datetime(owrd["calendar_month"]).dt.month
    pod = (
        owrd.groupby(["calendar_year", "calendar_month"], as_index=False)["volume_m3"]
        .sum()
        .rename(columns={"volume_m3": "owrd_direct_pod_m3"})
    )
    out = city_m.merge(pod, on=["calendar_year", "calendar_month"], how="outer")
    out["residual_m3"] = out["well_meter_for_sew_m3"] - out["owrd_direct_pod_m3"]
    out["identity_status"] = "unresolved"
    out["note"] = (
        "Evidence comparison only. Close totals or correlation do not identify "
        "WELL METER FOR SEW as OWRD direct POD withdrawal."
    )
    return out.sort_values(["calendar_year", "calendar_month"])


def evaluate_gate(qa: pd.DataFrame, double_count: dict, equiv_ok: bool) -> dict:
    blocking = qa[qa["severity"].eq("blocking") & qa["status"].eq("FAIL")]
    reasons = []
    passed = True
    if not equiv_ok:
        passed = False
        reasons.append("primary meter-report representations are not materially equivalent")
    if not blocking.empty:
        passed = False
        reasons.extend(blocking["check_id"].tolist())
    if double_count.get("identified_parent_submeter_double_counting"):
        passed = False
        reasons.append("identified parent/submeter double counting")
    # Units and consumption-month semantics are source-stated and implemented.
    reasons_ok = [
        "City units taken from City note: 1 unit = 100 ft3 = 748 US gallons",
        "Primary meter month treated as consumption month per City note",
        "city_metered_water_service_m3 excludes SWR, WELL METER FOR SEW, bulk, Trailer City, Warehouse",
        "2026 not_observed_yet cells excluded from observed aggregates",
    ]
    return {
        "gate": "PASS" if passed else "FAIL",
        "blocking_reasons": reasons,
        "notes": reasons_ok,
        "unresolved_excluded_from_response": [
            "swr_meter",
            "well_meter_for_sew",
            "bulk_water",
            "trailer_city_water",
            "warehouse_water",
        ],
        "double_count_investigation": {
            k: v
            for k, v in double_count.items()
            if k != "swap_month_overlaps"
        },
        "swap_month_overlaps": double_count.get("swap_month_overlaps", []),
    }


def _inventory_rows(hashes: dict) -> pd.DataFrame:
    # (source_id, scientific_role, intended_use, counts_as_scientific_source, canonical_filename)
    roles = {
        PRIMARY_XLSX: (
            SOURCE_ID_METER,
            "primary_structured",
            "Preferred machine-readable Facebook water/sewer meter report",
            True,
            PRIMARY_XLSX,
        ),
        PRIMARY_CSV: (
            SOURCE_ID_METER,
            "mirror_evidence",
            "CSV delivery of the same meter report; not an independent observation",
            True,
            PRIMARY_CSV,
        ),
        PRIMARY_CSV_DUP: (
            SOURCE_ID_METER,
            "duplicate_non_source",
            (
                "Byte-identical extra copy of FB Meters and Consumption(1).csv. "
                "Not an independently informative observation. Filesystem mtime "
                "predates the rest of the unpacked package. Classified as a local "
                "duplicate/non-source artifact created during handling. Preserved "
                "for raw immutability; not counted as a separate scientific source."
            ),
            False,
            PRIMARY_CSV,
        ),
        PRIMARY_TXT: (
            SOURCE_ID_METER,
            "mirror_evidence",
            "TXT delivery; CRLF-normalized content matches the CSV",
            True,
            PRIMARY_TXT,
        ),
        PRIMARY_PDF: (
            SOURCE_ID_METER,
            "mirror_evidence",
            "PDF print/evidence copy of the meter report",
            True,
            PRIMARY_PDF,
        ),
        BULK_XLSX: (
            SOURCE_ID_BULK,
            "primary_structured",
            "Facebook bulk/hydrant-water billing records",
            True,
            BULK_XLSX,
        ),
        EVENTS_XLSX: (
            SOURCE_ID_EVENTS,
            "primary_structured",
            "Meter set/swap/pull history compiled by the City from 2018 onward",
            True,
            EVENTS_XLSX,
        ),
        NOTE_TXT: (
            SOURCE_ID_NOTE,
            "source_metadata",
            "City explanatory note on units, read/bill timing, and lifecycle limits",
            True,
            NOTE_TXT,
        ),
    }
    rows = []
    for name, digest in hashes.items():
        sid, role, intent, counts, canonical = roles.get(
            name, ("UNREGISTERED", "unknown", "", False, name)
        )
        p = RAW / name
        rows.append(
            {
                "filename": name,
                "sha256": digest,
                "nbytes": p.stat().st_size if p.exists() else None,
                "source_id": sid,
                "custodian": "City of Prineville",
                "acquisition_package": "city_prineville_public_records_2026",
                "raw_format": p.suffix.lower().lstrip("."),
                "scientific_role": role,
                "intended_use": intent,
                "immutable": True,
                "counts_as_scientific_source": counts,
                "canonical_filename": canonical,
                "duplicate_of": None if name == canonical else canonical,
            }
        )
    return pd.DataFrame(rows)


def plot_observational_figures(
    components: pd.DataFrame,
    meters: pd.DataFrame,
    events: pd.DataFrame,
    recon: pd.DataFrame,
    owrd: pd.DataFrame,
) -> list[Path]:
    import matplotlib.pyplot as plt
    from matplotlib.dates import DateFormatter

    FIG.mkdir(parents=True, exist_ok=True)
    written = []

    def month_index(df, ycol="year", mcol="month"):
        return pd.to_datetime(dict(year=df[ycol].astype(int), month=df[mcol].astype(int), day=1))

    # 1. component time series
    fig, ax = plt.subplots(figsize=(11, 5.2))
    t = month_index(components)
    ax.plot(t, components["city_metered_water_service_m3"], label="City-metered service (observed)", lw=1.8)
    ax.plot(t, components["bulk_water_bill_month_m3"], label="Bulk water (bill month, unresolved)", lw=1.2, ls="--")
    ax.plot(t, components["swr_meter_volume_m3"], label="SWR METER (unresolved direction/identity)", lw=1.0, alpha=0.8)
    ax.plot(t, components["well_meter_for_sew_volume_m3"], label="WELL METER FOR SEW (unresolved)", lw=1.0, alpha=0.8)
    ax.set_ylabel("m³ / month")
    ax.set_title("City of Prineville meter components (not a campus water balance)")
    ax.legend(fontsize=8, loc="upper left")
    ax.text(
        0.01,
        -0.18,
        "Observed City records. Components are independent. They are not summed to total Meta withdrawal.",
        transform=ax.transAxes,
        fontsize=8,
    )
    fig.tight_layout()
    p = FIG / "city_water_components_monthly.png"
    fig.savefig(p, dpi=140)
    plt.close(fig)
    written.append(p)

    # 2. meter heatmap (Facebook Data Center only)
    dc = meters[meters["entity_name"].eq(ENTITY_FACEBOOK_DC)].copy()
    dc["period"] = dc["year"] * 100 + dc["month"]
    use = dc[dc["observation_status"].isin(["observed", "observed_zero"])]
    meters_order = sorted(use["meter_id_raw"].unique())
    periods = sorted(use["period"].unique())
    mat = np.full((len(meters_order), len(periods)), np.nan)
    idx_m = {m: i for i, m in enumerate(meters_order)}
    idx_p = {p0: j for j, p0 in enumerate(periods)}
    for r in use.itertuples(index=False):
        val = r.usage_m3
        if r.observation_status == "observed_zero":
            val = 0.0
        mat[idx_m[r.meter_id_raw], idx_p[int(r.period)]] = val
    fig, ax = plt.subplots(figsize=(12, max(4.5, 0.22 * len(meters_order))))
    im = ax.imshow(np.log1p(np.nan_to_num(mat, nan=np.nan)), aspect="auto", interpolation="nearest", cmap="viridis")
    ax.set_yticks(range(len(meters_order)))
    ax.set_yticklabels(meters_order, fontsize=7)
    tick_i = [i for i, p0 in enumerate(periods) if p0 % 100 == 1]
    ax.set_xticks(tick_i)
    ax.set_xticklabels([str(periods[i] // 100) for i in tick_i], fontsize=8)
    ax.set_title("Facebook Data Center meter activity (log1p m³); observed months only")
    fig.colorbar(im, ax=ax, fraction=0.02, label="log1p m³")
    fig.tight_layout()
    p = FIG / "city_meter_heatmap.png"
    fig.savefig(p, dpi=140)
    plt.close(fig)
    written.append(p)

    # 3. lifecycle timeline
    fig, ax = plt.subplots(figsize=(11, 4.8))
    first = (
        use.groupby("meter_id_raw")
        .apply(lambda g: pd.Timestamp(int(g.year.min()), int(g.loc[g.year.idxmin(), "month"] if False else g.month.min()), 1), include_groups=False)
    )
    # simpler first/last activity
    activity = use.groupby("meter_id_raw").agg(y0=("year", "min"), y1=("year", "max"))
    y_pos = {m: i for i, m in enumerate(sorted(activity.index))}
    for mid, r in activity.iterrows():
        ax.hlines(y_pos[mid], r.y0, r.y1 + 0.9, color="0.65", lw=2)
    for r in events.itertuples(index=False):
        if not r.event_date:
            continue
        y = int(r.event_date[:4]) + (int(r.event_date[5:7]) - 1) / 12
        act = (r.action or "").lower()
        color = "#2ca02c" if "set" in act and "swap" not in act else ("#d62728" if "pull" in act else "#ff7f0e")
        ax.axvline(y, color=color, alpha=0.25, lw=1)
    ax.axvline(2023 + 8 / 12, color="#9467bd", ls=":", lw=1.2, label="PRN1 chilled-water test 2023-09 (permit)")
    ax.axvline(2024 + 1 / 12, color="#8c564b", ls=":", lw=1.2, label="PRN1 chiller operational 2024-02 (permit)")
    ax.set_yticks(list(y_pos.values()))
    ax.set_yticklabels(list(y_pos.keys()), fontsize=7)
    ax.set_xlabel("year")
    ax.set_title("Meter activity span with City set/swap/pull events (events incomplete before 2018)")
    ax.legend(fontsize=7, loc="upper left")
    fig.tight_layout()
    p = FIG / "city_meter_lifecycle.png"
    fig.savefig(p, dpi=140)
    plt.close(fig)
    written.append(p)

    # 4. annual reconciliation
    fig, ax = plt.subplots(figsize=(10, 5))
    rr = recon.dropna(subset=["meta_annual_withdrawal_m3"], how="all")
    ax.plot(rr["year"], rr["meta_annual_withdrawal_m3"], "o-", label="Meta annual withdrawal (reported)", lw=2)
    ax.plot(rr["year"], rr["city_metered_water_service_m3"], "s--", label="City-metered service (provisional)")
    ax.plot(rr["year"], rr["bulk_water_bill_month_m3"], "d:", label="Bulk water (bill-year sum, unresolved)")
    ax.plot(rr["year"], rr["diagnostic_service_plus_bulk_m3"], "x-", label="service + bulk (bulk allocated by observed bill year)")
    ax.set_ylabel("m³ / year")
    ax.set_title("Meta annual withdrawal vs City meter components (boundary reconciliation)")
    ax.legend(fontsize=8)
    ax.text(
        0.01,
        -0.2,
        "Accounting-date diagnostic, not a water-balance closure. Diagnostic sum is not a campus boundary.",
        transform=ax.transAxes,
        fontsize=8,
    )
    fig.tight_layout()
    p = FIG / "meta_city_annual_reconciliation.png"
    fig.savefig(p, dpi=140)
    plt.close(fig)
    written.append(p)

    # 6. OWRD vs WELL METER FOR SEW if overlap
    both = owrd.dropna(subset=["well_meter_for_sew_m3", "owrd_direct_pod_m3"])
    if len(both) >= 12:
        fig, ax = plt.subplots(figsize=(10, 4.8))
        t = pd.to_datetime(dict(year=both.calendar_year.astype(int), month=both.calendar_month.astype(int), day=1))
        ax.plot(t, both["well_meter_for_sew_m3"], label="City WELL METER FOR SEW (unresolved identity)")
        ax.plot(t, both["owrd_direct_pod_m3"], label="OWRD Vitesse/Facebook direct POD sum")
        ax.set_ylabel("m³ / month")
        ax.set_title("City WELL METER FOR SEW vs OWRD direct POD (evidence comparison, not identity)")
        ax.legend(fontsize=8)
        fig.tight_layout()
        p = FIG / "owrd_vs_well_meter_for_sew.png"
        fig.savefig(p, dpi=140)
        plt.close(fig)
        written.append(p)
    return written


def _update_local_manifest(hashes: dict) -> None:
    """Update gitignored local source_manifest if present. Do not create a new convention."""
    path = ROOT / "data" / "source_manifest.csv"
    if not path.exists():
        return
    df = pd.read_csv(path)
    # Refresh CITY_PUBLIC_RECORDS status.
    mask = df["source_id"].eq("CITY_PUBLIC_RECORDS")
    if mask.any():
        df.loc[mask, "status"] = (
            "meter/bulk/lifecycle package received 2026; municipal well production, ASR, "
            "and complete sewer/return series remain outstanding"
        )
        df.loc[mask, "role"] = (
            "request mechanism; meter-level Facebook utility records now ingested under CITY_FB_* sources"
        )
    new_rows = [
        {
            "source_id": SOURCE_ID_METER,
            "category": "City utility meters",
            "title": "City of Prineville Facebook water/sewer meter consumption report",
            "url": "https://www.cityofprineville.com/1294/Public-Records",
            "resolution": "meter × month",
            "coverage": "2012-12 through 2026-07 (2012 and 2026 partial)",
            "role": "observed City-metered customer-service volumes; not total Meta withdrawal",
            "authority": "official primary",
            "access": f"local immutable XLSX sha256={hashes.get(PRIMARY_XLSX, '')[:12]}…",
            "status": "acquired; XLSX primary, CSV/TXT/PDF mirrors",
            "manual_required": False,
        },
        {
            "source_id": SOURCE_ID_BULK,
            "category": "City utility meters",
            "title": "City of Prineville Facebook bulk/hydrant water",
            "url": "https://www.cityofprineville.com/1294/Public-Records",
            "resolution": "bill-date rows",
            "coverage": "2018-2026 billing records",
            "role": "hydrant-meter bulk water by City billing convention",
            "authority": "official primary",
            "access": f"local immutable XLSX sha256={hashes.get(BULK_XLSX, '')[:12]}…",
            "status": "acquired",
            "manual_required": False,
        },
        {
            "source_id": SOURCE_ID_EVENTS,
            "category": "City utility meters",
            "title": "City of Prineville Facebook meter set/swap/pull history",
            "url": "https://www.cityofprineville.com/1294/Public-Records",
            "resolution": "event",
            "coverage": "compiled from ~2018 onward; incomplete earlier",
            "role": "meter lifecycle; not consumption",
            "authority": "official primary",
            "access": f"local immutable XLSX sha256={hashes.get(EVENTS_XLSX, '')[:12]}…",
            "status": "acquired; pre-2018 incomplete by City note",
            "manual_required": False,
        },
        {
            "source_id": SOURCE_ID_NOTE,
            "category": "City utility meters",
            "title": "City of Prineville data-description note",
            "url": "https://www.cityofprineville.com/1294/Public-Records",
            "resolution": "source metadata",
            "coverage": "accompanies 2026 delivery",
            "role": "units, read/bill timing, lifecycle limitation",
            "authority": "official primary",
            "access": f"local immutable TXT sha256={hashes.get(NOTE_TXT, '')[:12]}…",
            "status": "acquired",
            "manual_required": False,
        },
    ]
    existing = set(df["source_id"].astype(str))
    add = pd.DataFrame([r for r in new_rows if r["source_id"] not in existing])
    if not add.empty:
        df = pd.concat([df, add], ignore_index=True)
    else:
        for r in new_rows:
            m = df["source_id"].eq(r["source_id"])
            if m.any():
                for k, v in r.items():
                    if k in df.columns:
                        df.loc[m, k] = v
    df.to_csv(path, index=False)

    pri = ROOT / "data" / "canonical" / "source_priority_matrix.csv"
    if pri.exists():
        p = pd.read_csv(pri)
        m = p["quantity"].eq("meta_monthly_water")
        if m.any():
            p.loc[m, "preferred_source"] = "City of Prineville Facebook Data Center WATER-COMM + ADD'L WATER meters"
            p.loc[m, "resolution"] = "month (consumption month per City note)"
            p.loc[m, "coverage"] = (
                "City-service observed 2012-12 through 2026-07 (2012 and 2026 partial); "
                "all-source monthly campus withdrawal unresolved"
            )
            p.loc[m, "role_if_available"] = (
                "observed City customer-service water component; NOT total Meta campus withdrawal"
            )
            p.loc[m, "fallback_if_missing"] = (
                "do not substitute the physics-shaped Meta-total monthly reconstruction for "
                "observed City-service; that reconstruction remains a model-generated "
                "proxy/scenario for unresolved all-source monthly campus withdrawal"
            )
            p.loc[m, "validation_status"] = (
                "City-service observed through 2026-07; complete campus monthly "
                "withdrawal/source mix/return balance unresolved"
            )
        p.to_csv(pri, index=False)


def run() -> dict:
    OUT.mkdir(parents=True, exist_ok=True)
    QC.mkdir(parents=True, exist_ok=True)
    if not RAW.exists():
        raise SystemExit(f"Missing City source directory: {RAW}")

    meta_hash_before = sha256_file(META_ANNUAL) if META_ANNUAL.exists() else ""
    hashes = {p.name: sha256_file(p) for p in sorted(RAW.iterdir()) if p.is_file()}
    inv = _inventory_rows(hashes)
    inv.to_csv(OUT / "city_source_file_inventory.csv", index=False)

    xlsx_path = RAW / PRIMARY_XLSX
    csv_path = RAW / PRIMARY_CSV
    txt_path = RAW / PRIMARY_TXT
    meters = parse_primary_meter_report(xlsx_path)
    csv_long = parse_primary_delimited(csv_path)
    txt_long = parse_primary_delimited(txt_path)
    equiv = compare_meter_representations(meters, csv_long)
    csv_bytes = csv_path.read_bytes().replace(b"\r\n", b"\n")
    txt_bytes = txt_path.read_bytes().replace(b"\r\n", b"\n")
    txt_vs_csv = {
        "csv_sha256": hashes.get(PRIMARY_CSV),
        "txt_sha256": hashes.get(PRIMARY_TXT),
        "extra_csv_byte_identical_to_csv1": (
            (RAW / PRIMARY_CSV_DUP).exists()
            and (RAW / PRIMARY_CSV_DUP).read_bytes() == csv_path.read_bytes()
        ),
        "equivalent_after_newline_normalize": csv_bytes == txt_bytes,
        "csv_minus_txt_bytes": len(csv_path.read_bytes()) - len(txt_path.read_bytes()),
    }
    txt_equiv = compare_meter_representations(meters, txt_long)
    equiv["txt_vs_xlsx_materially_equivalent"] = txt_equiv["materially_equivalent"]
    pdf_info = pdf_evidence_tokens(RAW / PRIMARY_PDF)

    bulk = parse_bulk_water(RAW / BULK_XLSX)
    events = parse_meter_events(RAW / EVENTS_XLSX)
    lineage = build_lineage_audit(events, meters)
    double_count = parent_submeter_investigation(meters, events)

    comps = components_monthly(meters, bulk)
    cov = coverage_table(meters, bulk)
    meta = pd.read_csv(META_ANNUAL) if META_ANNUAL.exists() else pd.DataFrame(columns=["year", "water_withdrawal_m3_reported"])
    recon = meta_reconciliation(comps, meta)
    owrd = owrd_well_meter_compare(meters, OWRD_DIRECT)

    meta_hash_after = sha256_file(META_ANNUAL) if META_ANNUAL.exists() else ""
    qa, exc = build_qa(
        inv,
        equiv,
        txt_vs_csv,
        pdf_info,
        meters,
        bulk,
        events,
        lineage,
        meta_hash_before,
        meta_hash_after,
    )
    gate = evaluate_gate(qa, double_count, equiv["materially_equivalent"] and txt_equiv["materially_equivalent"])

    # persist
    meters.to_csv(OUT / "city_meter_monthly_long.csv", index=False)
    bulk.to_csv(OUT / "city_bulk_water_monthly.csv", index=False)
    events.to_csv(OUT / "city_meter_events.csv", index=False)
    lineage.to_csv(OUT / "city_meter_lineage_audit.csv", index=False)
    qa.to_csv(OUT / "city_meter_qa.csv", index=False)
    exc.to_csv(OUT / "city_meter_qa_exceptions.csv", index=False)
    comps.to_csv(OUT / "city_water_components_monthly.csv", index=False)
    cov.to_csv(OUT / "city_source_coverage.csv", index=False)
    recon.to_csv(OUT / "city_meta_annual_reconciliation.csv", index=False)
    owrd.to_csv(OUT / "city_owrd_well_meter_compare.csv", index=False)
    qa.to_csv(QC / "city_prineville_utility_qa.csv", index=False)
    (OUT / "model_promotion_gate.json").write_text(json.dumps(gate, indent=2, default=str) + "\n")
    (OUT / "representation_equivalence.json").write_text(
        json.dumps({"xlsx_csv": equiv, "csv_txt": txt_vs_csv, "pdf": pdf_info}, indent=2) + "\n"
    )
    (OUT / "double_count_investigation.json").write_text(json.dumps(double_count, indent=2, default=str) + "\n")

    figs = plot_observational_figures(comps, meters, events, recon, owrd)
    _update_local_manifest(hashes)

    summary = {
        "n_meter_month": int(len(meters)),
        "n_unique_meters": int(meters["meter_id_raw"].nunique()),
        "entities": sorted(meters["entity_name"].unique().tolist()),
        "service_locations": sorted(meters["service_location"].unique().tolist()),
        "rate_codes": sorted(meters["rate_code_raw"].unique().tolist()),
        "years": [int(y) for y in sorted(meters["year"].unique())],
        "latest_observed": {
            "year": meters.attrs.get("latest_observed_year"),
            "month": meters.attrs.get("latest_observed_month"),
        },
        "n_bulk_rows": int(len(bulk)),
        "n_events": int(len(events)),
        "gate": gate["gate"],
        "figures": [str(p.relative_to(ROOT)) for p in figs],
        "meta_annual_sha256": meta_hash_after,
    }
    (OUT / "parser_summary.json").write_text(json.dumps(summary, indent=2) + "\n")
    print(json.dumps(summary, indent=2))
    return summary


if __name__ == "__main__":
    run()
